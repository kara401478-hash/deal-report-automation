"""
商談・案件実績 自動集計レポートスクリプト
社内システムから当月の商談・案件データをCSV取得し、
個人別集計Excelを自動生成する。
OneDriveへのコピーとOutlookによるメール送信まで完全自動化。
タスクスケジューラでの定期実行を想定。
"""

import os
import time
import shutil
import traceback
from datetime import datetime
from dotenv import load_dotenv

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

# ============================================================
# 設定読み込み
# ============================================================
load_dotenv()

LOGIN_ID             = os.getenv("LOGIN_ID")
LOGIN_PASSWORD       = os.getenv("LOGIN_PASSWORD")
TARGET_URL           = os.getenv("TARGET_URL")
NAV_LABEL            = os.getenv("NAV_LABEL")
NAV_SUBLABEL_DEAL    = os.getenv("NAV_SUBLABEL_DEAL")
NAV_SUBLABEL_PROJECT = os.getenv("NAV_SUBLABEL_PROJECT")
DOWNLOAD_FOLDER      = os.getenv("DOWNLOAD_FOLDER")
WORK_EXCEL_PATH      = os.getenv("WORK_EXCEL_PATH")
TABLE_EXCEL_PATH     = os.getenv("TABLE_EXCEL_PATH")
OUTPUT_EXCEL_PATH    = os.getenv("OUTPUT_EXCEL_PATH")
ONEDRIVE_PATH        = os.getenv("ONEDRIVE_PATH")
MAIL_FROM            = os.getenv("MAIL_FROM")
MAIL_TO              = os.getenv("MAIL_TO")
MAIL_CC              = os.getenv("MAIL_CC")
MAIL_SUBJECT         = os.getenv("MAIL_SUBJECT")
ONEDRIVE_URL         = os.getenv("ONEDRIVE_URL")

# ============================================================
# ユーティリティ関数
# ============================================================

def setup_driver(download_folder):
    """ChromeDriverのセットアップ"""
    options = Options()
    prefs = {
        "download.default_directory": download_folder,
        "download.prompt_for_download": False,
        "directory_upgrade": True,
        "safebrowsing.enabled": True,
    }
    options.add_experimental_option("prefs", prefs)
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    return driver


def wait_for_downloads(download_folder, timeout=60):
    """ダウンロード完了まで待機"""
    for _ in range(timeout):
        time.sleep(1)
        if not any(f.endswith(".crdownload") for f in os.listdir(download_folder)):
            print("Download complete.")
            return True
    print("Timeout: Download did not complete.")
    return False


def rename_latest_csv(download_folder, new_file_path):
    """最新のCSVを指定パスにリネーム"""
    csv_files = [f for f in os.listdir(download_folder) if f.endswith(".csv")]
    if not csv_files:
        print("No CSV files found.")
        return
    latest = max(
        [os.path.join(download_folder, f) for f in csv_files],
        key=os.path.getctime
    )
    if os.path.exists(new_file_path):
        os.remove(new_file_path)
    os.rename(latest, new_file_path)
    print(f"Saved: {new_file_path}")


def set_date_input(driver, wait, xpath, date_str):
    """日付入力欄に直接テキストを入力"""
    el = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
    el.click()
    el.send_keys(Keys.CONTROL + "a")
    el.send_keys(date_str)
    el.send_keys(Keys.TAB)
    time.sleep(0.5)


# ============================================================
# Step1: CSVダウンロード
# ============================================================

def download_data(driver, wait):
    """商談・案件データをCSVでダウンロード"""
    today = datetime.today()
    first_day_str = today.replace(day=1).strftime("%Y/%m/%d")
    today_str = today.strftime("%Y/%m/%d")
    print(f"集計期間: {first_day_str} ～ {today_str}")

    # ログイン
    driver.get(TARGET_URL)
    wait.until(EC.presence_of_element_located(
        (By.XPATH, '/html/body/div/form/div[1]/input')
    )).send_keys(LOGIN_ID)
    driver.find_element(By.XPATH, '/html/body/div/form/div[2]/input').send_keys(LOGIN_PASSWORD)
    wait.until(EC.element_to_be_clickable(
        (By.XPATH, '/html/body/div/form/input[2]')
    )).click()
    print("✅ ログイン完了")

    # 商談CSV取得
    deal_csv = os.path.join(DOWNLOAD_FOLDER, "1.deals.csv")
    try:
        wait.until(EC.element_to_be_clickable(
            (By.XPATH, f"//span[contains(text(),'{NAV_LABEL}')]")
        )).click()
        wait.until(EC.element_to_be_clickable(
            (By.XPATH, f"//span[@class='nav-label' and text()='{NAV_SUBLABEL_DEAL}']")
        )).click()
        wait.until(EC.element_to_be_clickable(
            (By.XPATH, '/html/body/div[2]/div/div[3]/div/div/div/div[1]/div/button')
        )).click()

        set_date_input(driver, wait,
            "/html/body/div[2]/div/div[3]/div/div/div/div[1]/div/div[2]/div/div/form/div[2]/div[21]/div[2]/div/input[1]",
            first_day_str)
        set_date_input(driver, wait,
            "/html/body/div[2]/div/div[3]/div/div/div/div[1]/div/div[2]/div/div/form/div[2]/div[21]/div[2]/div/input[2]",
            today_str)

        wait.until(EC.element_to_be_clickable(
            (By.XPATH, '/html/body/div[2]/div/div[3]/div/div/div/div[1]/div/div[2]/div/div/form/div[3]/input')
        )).click()
        time.sleep(10)

        btn = wait.until(EC.element_to_be_clickable(
            (By.XPATH, '/html/body/div[2]/div/div[3]/div/div/div/div[2]/div[1]/div[1]/div/button')
        ))
        driver.execute_script("arguments[0].scrollIntoView(true);", btn)
        driver.execute_script("arguments[0].click();", btn)

        wait.until(EC.element_to_be_clickable(
            (By.XPATH, '/html/body/div[2]/div/div[3]/div/div/div/div[2]/div[1]/div[1]/div/ul/li[1]/a')
        )).click()
        time.sleep(30)

        WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//span[text()="CSVダウンロード"]'))
        ).click()
        wait.until(EC.element_to_be_clickable(
            (By.XPATH, '/html/body/div[2]/div/div[3]/div/div/div/div[2]/div[2]/table/tbody/tr[1]/td[1]/ul/li[1]/a')
        )).click()

        wait_for_downloads(DOWNLOAD_FOLDER)
        rename_latest_csv(DOWNLOAD_FOLDER, deal_csv)
        print("✅ 商談CSV取得完了")
    except Exception as e:
        print(f"❌ 商談処理エラー: {e}")

    # 案件CSV取得
    project_csv = os.path.join(DOWNLOAD_FOLDER, "2.projects.csv")
    try:
        wait.until(EC.element_to_be_clickable(
            (By.XPATH, f"//span[contains(text(),'{NAV_LABEL}')]")
        )).click()
        wait.until(EC.element_to_be_clickable(
            (By.XPATH, f"//span[@class='nav-label' and text()='{NAV_SUBLABEL_PROJECT}']")
        )).click()
        wait.until(EC.element_to_be_clickable(
            (By.XPATH, '/html/body/div[2]/div/div[3]/div/div/div/div[1]/div/button')
        )).click()

        set_date_input(driver, wait,
            "/html/body/div[2]/div/div[3]/div/div/div/div[1]/div/div[2]/div/div/form/div[2]/div[245]/div[2]/div/input[1]",
            first_day_str)
        set_date_input(driver, wait,
            "/html/body/div[2]/div/div[3]/div/div/div/div[1]/div/div[2]/div/div/form/div[2]/div[245]/div[2]/div/input[2]",
            today_str)

        wait.until(EC.element_to_be_clickable(
            (By.XPATH, '/html/body/div[2]/div/div[3]/div/div/div/div[1]/div/div[2]/div/div/form/div[3]/input')
        )).click()
        time.sleep(10)

        btn = wait.until(EC.element_to_be_clickable(
            (By.XPATH, '/html/body/div[2]/div/div[3]/div/div/div/div[2]/div[1]/div[1]/div/button')
        ))
        driver.execute_script("arguments[0].scrollIntoView(true);", btn)
        driver.execute_script("arguments[0].click();", btn)

        for item in wait.until(EC.presence_of_all_elements_located((By.XPATH, '//ul/li/a'))):
            if item.text.strip() == "全項目":
                driver.execute_script("arguments[0].click();", item)
                break
        time.sleep(30)

        WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//span[text()="CSVダウンロード"]'))
        ).click()
        wait.until(EC.element_to_be_clickable(
            (By.XPATH, '/html/body/div[2]/div/div[3]/div/div/div/div[2]/div[2]/table/tbody/tr[1]/td[1]/ul/li[1]/a')
        )).click()

        wait_for_downloads(DOWNLOAD_FOLDER)
        rename_latest_csv(DOWNLOAD_FOLDER, project_csv)
        print("✅ 案件CSV取得完了")
    except Exception as e:
        print(f"❌ 案件処理エラー: {e}")


# ============================================================
# Step2: Excel生成・集計シート作成
# ============================================================

def build_work_excel():
    """CSVをExcelに取り込み、変換シート・集計シートを生成"""
    deal_csv    = os.path.join(DOWNLOAD_FOLDER, "1.deals.csv")
    project_csv = os.path.join(DOWNLOAD_FOLDER, "2.projects.csv")

    deal_df    = pd.read_csv(deal_csv,    encoding="cp932")
    project_df = pd.read_csv(project_csv, encoding="cp932")
    table_sheets = pd.read_excel(TABLE_EXCEL_PATH, sheet_name=None)

    with pd.ExcelWriter(WORK_EXCEL_PATH, mode="w") as writer:
        deal_df.to_excel(writer,    sheet_name="deals",    index=False)
        project_df.to_excel(writer, sheet_name="projects", index=False)
        for sheet_name, df in table_sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    print(f"✅ 作業用Excel保存: {WORK_EXCEL_PATH}")

    # 日付列の整形
    wb = openpyxl.load_workbook(WORK_EXCEL_PATH)
    ws = wb["deals"]
    for row in range(2, ws.max_row + 1):
        for col in [30, 31]:
            cell = ws.cell(row=row, column=col)
            if cell.value:
                try:
                    dt = pd.to_datetime(str(cell.value), errors="coerce")
                    if pd.notna(dt):
                        cell.value = dt.strftime("%Y/%m/%d")
                except Exception:
                    pass

    # 変換シート（deals_converted）
    _add_deals_converted_sheet(wb)

    # 変換シート（projects_converted）
    _add_projects_converted_sheet(wb)

    # 個人別集計シート
    _add_summary_sheet(wb)

    wb.save(WORK_EXCEL_PATH)
    print(f"✅ 集計シート生成完了: {WORK_EXCEL_PATH}")


def _add_deals_converted_sheet(wb):
    sheet_name = "deals_converted"
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    ws = wb[sheet_name]

    headers = ["システムID", "商談日付", "商談タイトル", "商談区分", "新規", "休眠", "担当者"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)

    src = wb["deals"]
    unique_ids = list(dict.fromkeys(
        cell.value for row in src.iter_rows(min_row=2, min_col=1, max_col=1)
        for cell in row if cell.value
    ))
    for i, uid in enumerate(unique_ids, start=2):
        ws[f"A{i}"] = uid
        ws[f"B{i}"] = f'=IFERROR(TEXT(EOMONTH(INDEX(deals!M:M,MATCH(deals_converted!A{i},deals!A:A,0)),-1)+1,"yyyy/mm/dd")," ")'
        ws[f"C{i}"] = f'=IFERROR(INDEX(deals!P:P,MATCH(deals_converted!A{i},deals!A:A,0))," ")'
        ws[f"D{i}"] = f'=IFERROR(IF(VLOOKUP($C{i},deal_table!$A:$D,2,0)=0," ",VLOOKUP(deals_converted!$C{i},deal_table!$A:$D,2,0))," ")'
        ws[f"E{i}"] = f'=IFERROR(IF(VLOOKUP($C{i},deal_table!$A:$D,3,0)=0," ",VLOOKUP(deals_converted!$C{i},deal_table!$A:$D,3,0))," ")'
        ws[f"F{i}"] = f'=IFERROR(IF(VLOOKUP($C{i},deal_table!$A:$D,4,0)=0," ",VLOOKUP(deals_converted!$C{i},deal_table!$A:$D,4,0))," ")'
        ws[f"G{i}"] = f'=IFERROR(INDEX(deals!W:W,MATCH(deals_converted!A{i},deals!A:A,0))," ")'

    ws.freeze_panes = "A2"
    for col in ["A","B","C","D","E","F","G"]:
        ws.column_dimensions[col].width = 15


def _add_projects_converted_sheet(wb):
    sheet_name = "projects_converted"
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    ws = wb[sheet_name]

    headers = ["システムID", "案件種別", "営業担当", "作成日"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)

    src = wb["projects"]
    unique_ids = list(dict.fromkeys(
        cell.value for row in src.iter_rows(min_row=2, min_col=1, max_col=1)
        for cell in row if cell.value
    ))
    for i, uid in enumerate(unique_ids, start=2):
        ws[f"A{i}"] = uid
        ws[f"B{i}"] = f'=IFERROR(VLOOKUP(INDEX(projects!J:J,MATCH(projects_converted!A{i},projects!A:A,0)),project_table!$A:$B,2,0)," ")'
        ws[f"C{i}"] = f'=IFERROR(INDEX(projects!N:N,MATCH(projects_converted!A{i},projects!A:A,0))," ")'
        ws[f"D{i}"] = f'=IFERROR(TEXT(EOMONTH(DATEVALUE(LEFT(INDEX(projects!CI:CI,MATCH(projects_converted!A{i},projects!A:A,0)),10)),-1)+1,"yyyy/mm/dd")," ")'

    ws.freeze_panes = "A2"
    for col in ["A","B","C","D"]:
        ws.column_dimensions[col].width = 15


def _add_summary_sheet(wb):
    sheet_name = "individual_summary"
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name, index=0)
    ws = wb[sheet_name]

    ws["O1"] = '=TEXT(TODAY()-1,"yyyy/mm/dd")&"現在"'
    ws["A1"] = '=TEXT(IF(O1="","",EOMONTH(DATEVALUE(LEFT(O1,10)),-1)+1),"yyyy/mm/dd")'

    for cell_range, label in [("D2:F2","期間計"),("G2:I2","（新規）"),("J2:L2","（休眠）"),("M2:O2","（既存）")]:
        ws[cell_range.split(":")[0]] = label
        ws.merge_cells(cell_range)
        ws[cell_range.split(":")[0]].alignment = Alignment(horizontal="center")

    row3_headers = ["部署","所属","担当者","商談数","案件数","受注率","商談数","案件数","受注率","商談数","案件数","受注率","商談数","案件数","受注率"]
    for col, h in enumerate(row3_headers, start=1):
        ws.cell(row=3, column=col, value=h)

    src = wb["deals"]
    unique_reps = list(dict.fromkeys(
        cell.value for row in src.iter_rows(min_row=2, min_col=23, max_col=23)
        for cell in row if cell.value
    ))
    for i, rep in enumerate(unique_reps, start=4):
        ws[f"A{i}"] = f'=IFERROR(INDEX(deals!U:U,MATCH(individual_summary!$C{i},deals!$W:$W,0)),"-")'
        ws[f"B{i}"] = f'=IFERROR(INDEX(deals!V:V,MATCH(individual_summary!$C{i},deals!$W:$W,0)),"-")'
        ws[f"C{i}"] = rep
        ws[f"D{i}"] = f'=IFERROR(COUNTIFS(deals_converted!B:B,individual_summary!$A$1,deals_converted!D:D,deals_converted!$D$1,deals_converted!G:G,individual_summary!$C{i}),"-")'
        ws[f"E{i}"] = f'=IFERROR(COUNTIFS(projects_converted!D:D,individual_summary!$A$1,projects_converted!C:C,individual_summary!$C{i}),"-")'
        ws[f"F{i}"] = f'=IFERROR(E{i}/D{i},"-")'
        ws[f"G{i}"] = f'=IFERROR(COUNTIFS(deals_converted!B:B,individual_summary!$A$1,deals_converted!D:D,deals_converted!$D$1,deals_converted!G:G,individual_summary!$C{i},deals_converted!E:E,deals_converted!$E$1),"-")'
        ws[f"H{i}"] = f'=IFERROR(COUNTIFS(projects_converted!D:D,individual_summary!$A$1,projects_converted!C:C,individual_summary!$C{i},projects_converted!B:B,deals_converted!$E$1),"-")'
        ws[f"I{i}"] = f'=IFERROR(H{i}/G{i},"-")'
        ws[f"J{i}"] = f'=IFERROR(COUNTIFS(deals_converted!B:B,individual_summary!$A$1,deals_converted!D:D,deals_converted!$D$1,deals_converted!G:G,individual_summary!$C{i},deals_converted!F:F,deals_converted!$F$1),"-")'
        ws[f"K{i}"] = f'=IFERROR(COUNTIFS(projects_converted!D:D,individual_summary!$A$1,projects_converted!C:C,individual_summary!$C{i},projects_converted!B:B,deals_converted!$F$1),"-")'
        ws[f"L{i}"] = f'=IFERROR(K{i}/J{i},"-")'
        ws[f"M{i}"] = f'=IF(OR(ISBLANK(D{i}),ISBLANK(G{i}),ISBLANK(J{i})),"-",D{i}-G{i}-J{i})'
        ws[f"N{i}"] = f'=IFERROR(COUNTIFS(projects_converted!D:D,individual_summary!$A$1,projects_converted!C:C,individual_summary!$C{i},projects_converted!B:B,"existing"),"-")'
        ws[f"O{i}"] = f'=IFERROR(N{i}/M{i},"-")'
        for col in ["F","I","L","O"]:
            ws[f"{col}{i}"].number_format = "0.0%"

    ws.freeze_panes = "A4"
    for col in ["A","B","C"]:
        ws.column_dimensions[col].width = 15
    for col in ["D","E","F","G","H","I","J","K","L","M","N","O"]:
        ws.column_dimensions[col].width = 10


# ============================================================
# Step3: 最終Excel出力（書式整形）
# ============================================================

def build_output_excel():
    """集計シートを計算値に変換し書式整形して最終Excelとして出力"""
    import xlwings as xw
    from xlwings.constants import LineStyle

    app = xw.App(visible=False)
    try:
        wb = app.books.open(WORK_EXCEL_PATH)
        src = wb.sheets["individual_summary"]

        new_wb = xw.Book()
        new_ws = new_wb.sheets[0]
        new_ws.name = "individual_summary"
        new_ws.range("A1").value = src.used_range.value

        for cell_range in ["D2:F2","G2:I2","J2:L2","M2:O2"]:
            new_ws.range(cell_range).merge()
            new_ws.range(cell_range.split(":")[0]).api.HorizontalAlignment = -4108

        for col in ["F","I","L","O"]:
            new_ws.range(f"{col}:{col}").number_format = "0.0%"

        for col in ["A","B","C"]:
            new_ws.range(f"{col}:{col}").column_width = 15
        for col in ["D","E","F","G","H","I","J","K","L","M","N","O"]:
            new_ws.range(f"{col}:{col}").column_width = 10

        last_row = new_ws.range("A" + str(new_ws.cells.last_cell.row)).end("up").row
        new_ws.range(f"A3:O{last_row}").api.Borders.LineStyle = LineStyle.xlContinuous
        new_ws.range(f"A3:O{last_row}").api.Borders.Weight = 2

        new_ws.api.Activate()
        new_ws.api.Range("A4").Select()
        app.api.ActiveWindow.FreezePanes = True
        new_ws.range("A3:O3").api.AutoFilter(1)

        wb.close()
        new_wb.save(OUTPUT_EXCEL_PATH)
        print(f"✅ 出力Excel保存完了: {OUTPUT_EXCEL_PATH}")
    except Exception as e:
        print(f"❌ 出力Excel生成エラー: {e}")
        traceback.print_exc()
    finally:
        new_wb.close()
        app.quit()


# ============================================================
# Step4: OneDriveコピー & メール送信
# ============================================================

def copy_to_onedrive():
    """OneDriveにファイルをコピー"""
    dest_dir = os.path.dirname(ONEDRIVE_PATH)
    if not os.path.exists(dest_dir):
        raise FileNotFoundError(f"コピー先ディレクトリが見つかりません: {dest_dir}")
    if os.path.exists(ONEDRIVE_PATH):
        os.remove(ONEDRIVE_PATH)
    shutil.copy2(OUTPUT_EXCEL_PATH, ONEDRIVE_PATH)
    print(f"✅ OneDriveコピー完了: {ONEDRIVE_PATH}")


def send_mail():
    """Outlookでメール送信"""
    import win32com.client as win32
    outlook = win32.Dispatch("Outlook.Application")

    accounts = outlook.Session.Accounts
    target_account = next(
        (a for a in accounts if a.SmtpAddress.lower() == MAIL_FROM.lower()), None
    )

    mail = outlook.CreateItem(0)
    if target_account:
        mail.SendUsingAccount = target_account
    mail.To = MAIL_TO
    mail.CC = MAIL_CC
    mail.Subject = MAIL_SUBJECT
    mail.HTMLBody = f"""<html><body>
<p>各位</p>
<p>{MAIL_SUBJECT}の集計をお送りします。</p>
<p><a href="{ONEDRIVE_URL}">レポートを開く</a></p>
<p>以上、ご確認ください。</p>
</body></html>"""
    mail.Send()
    print(f"✅ メール送信完了: {datetime.now()}")


# ============================================================
# メイン処理
# ============================================================

def main():
    # Step1: CSVダウンロード
    driver = setup_driver(DOWNLOAD_FOLDER)
    wait = WebDriverWait(driver, 20)
    try:
        download_data(driver, wait)
    except Exception as e:
        print(f"❌ ダウンロードエラー: {e}")
        traceback.print_exc()
    finally:
        driver.quit()

    # Step2: Excel生成
    build_work_excel()

    # Step3: 最終Excel出力
    build_output_excel()

    # Step4: OneDriveコピー & メール送信
    try:
        copy_to_onedrive()
        send_mail()
    except Exception as e:
        print(f"❌ コピー/メール送信エラー: {e}")
        traceback.print_exc()


if __name__ == "__main__":
    main()
